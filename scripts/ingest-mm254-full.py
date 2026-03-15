#!/usr/bin/env python3
"""
Full ingestion of Mall Mediano (MM254) data into pg-arauco.
Replicates the MG446 pipeline for the MM building.

Steps:
  1. CSV → raw_readings (8.9M rows)
  2. CSV → store (254 meters with real store_types)
  3. raw_readings → meter_monthly (aggregated)
  4. raw_readings → meter_readings (partitioned)
  5. XLSX Resumen Mensual → meter_monthly_billing
  6. XLSX Resumen Mensual → meter_monthly_billing KPIs (update)
  7. XLSX Pliegos Tarifarios (Santiago) → tariff (with location)
  8. XLSX Resumen Mensual → building_summary (Arauco Estación)

Usage: python3 scripts/ingest-mm254-full.py
"""
import subprocess, sys, os, csv, io, time
for pkg in ['psycopg2-binary', 'openpyxl']:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

import psycopg2
import openpyxl
from datetime import date

# --- Config ---
CSV_PATH = 'docs/data/MALL_MEDIANO_254_completo.csv'
XLSX_PATH = 'docs/data/MM254_KPIs_mensuales_2025_M.xlsx'
BUILDING = 'Arauco Estación'
YEAR = 2025
METER_PREFIX = 'MM'
BATCH_SIZE = 5000

MONTH_MAP = {
    'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
    'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
    'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12,
}

conn = psycopg2.connect(host='127.0.0.1', port=5434, dbname='arauco', user='postgres', password='arauco')
cur = conn.cursor()

def parse_decimal(val):
    """Parse comma-decimal string to float, or None."""
    if val is None or val == '':
        return None
    return float(val.replace(',', '.'))


# =============================================
# STEP 1: CSV → raw_readings
# =============================================
print('=' * 60)
print('STEP 1: CSV → raw_readings')
print('=' * 60)

# Check if already loaded
cur.execute("SELECT COUNT(*) FROM raw_readings WHERE meter_id LIKE 'MM-%'")
existing = cur.fetchone()[0]
if existing > 0:
    print(f'  Ya existen {existing:,} filas MM en raw_readings, saltando...')
else:
    start = time.time()
    total = 0
    batch = []

    with open(CSV_PATH, 'r', encoding='latin1') as f:
        reader = csv.reader(f, delimiter=';')
        header = next(reader)  # skip header
        # Columns: 0:timestamp, 1:meter_id, 16:power_kW, 17:reactive_power_kvar, 18:power_factor, 20:energy_kWh_total
        for row in reader:
            batch.append((
                row[0],                    # timestamp
                row[1],                    # meter_id
                parse_decimal(row[16]),    # power_kw
                parse_decimal(row[17]),    # reactive_power_kvar
                parse_decimal(row[18]),    # power_factor
                parse_decimal(row[20]),    # energy_kwh_total
            ))
            if len(batch) >= BATCH_SIZE:
                args = ','.join(
                    cur.mogrify("(%s,%s,%s,%s,%s,%s)", b).decode()
                    for b in batch
                )
                cur.execute(f"""
                    INSERT INTO raw_readings (timestamp, meter_id, power_kw, reactive_power_kvar, power_factor, energy_kwh_total)
                    VALUES {args}
                    ON CONFLICT DO NOTHING
                """)
                total += len(batch)
                batch = []
                if total % 500000 == 0:
                    conn.commit()
                    elapsed = time.time() - start
                    rate = total / elapsed if elapsed > 0 else 0
                    print(f'  {total:,} filas | {rate:,.0f} filas/s | {elapsed:.0f}s')

    if batch:
        args = ','.join(
            cur.mogrify("(%s,%s,%s,%s,%s,%s)", b).decode()
            for b in batch
        )
        cur.execute(f"""
            INSERT INTO raw_readings (timestamp, meter_id, power_kw, reactive_power_kvar, power_factor, energy_kwh_total)
            VALUES {args}
            ON CONFLICT DO NOTHING
        """)
        total += len(batch)

    conn.commit()
    elapsed = time.time() - start
    print(f'  Total raw_readings insertados: {total:,} en {elapsed:.1f}s')

cur.execute("SELECT COUNT(*) FROM raw_readings WHERE meter_id LIKE 'MM-%'")
print(f'  Verificación raw_readings MM: {cur.fetchone()[0]:,}')


# =============================================
# STEP 2: CSV → store (254 meters with real store_types)
# =============================================
print('\n' + '=' * 60)
print('STEP 2: CSV → store (254 medidores)')
print('=' * 60)

# Extract unique (meter_id, store_type, store_name) from CSV — early exit once all found
store_info = {}  # meter_id → (store_type_name, store_name)
with open(CSV_PATH, 'r', encoding='latin1') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)  # skip header
    for row in reader:
        mid = row[1]
        if mid not in store_info:
            store_info[mid] = (row[4], row[5])  # store_type, store_name
            if len(store_info) >= 254:
                break

print(f'  Medidores únicos en CSV: {len(store_info)}')

# Ensure all store_types exist
unique_store_types = set(st for st, _ in store_info.values())
print(f'  Tipos de local únicos: {len(unique_store_types)}')

store_type_map = {}  # name → id
cur.execute("SELECT COALESCE(MAX(id), 0) FROM store_type")
next_id = cur.fetchone()[0] + 1

for st_name in sorted(unique_store_types):
    cur.execute("SELECT id FROM store_type WHERE name = %s", (st_name,))
    row = cur.fetchone()
    if row:
        store_type_map[st_name] = row[0]
    else:
        cur.execute("INSERT INTO store_type (id, name) VALUES (%s, %s)", (next_id, st_name))
        store_type_map[st_name] = next_id
        print(f'    Nuevo store_type: "{st_name}" → id={next_id}')
        next_id += 1

conn.commit()

# Insert stores (idempotent)
stores = []
for mid in sorted(store_info.keys()):
    st_name, s_name = store_info[mid]
    stores.append((mid, s_name, store_type_map[st_name]))

cur.executemany("""
    INSERT INTO store (meter_id, store_name, store_type_id) VALUES (%s, %s, %s)
    ON CONFLICT (meter_id) DO NOTHING
""", stores)
conn.commit()
print(f'  Stores insertados/existentes: {len(stores)}')


# =============================================
# STEP 3: raw_readings → meter_monthly
# =============================================
print('\n' + '=' * 60)
print('STEP 3: raw_readings → meter_monthly')
print('=' * 60)

cur.execute("DELETE FROM meter_monthly WHERE meter_id LIKE 'MM-%'")
print(f'  meter_monthly MM eliminados previos: {cur.rowcount}')
conn.commit()

SQL_AGG = """
INSERT INTO meter_monthly (meter_id, month, total_kwh, avg_power_kw, peak_power_kw, total_reactive_kvar, avg_power_factor)
SELECT
    meter_id,
    date_trunc('month', timestamp)::date AS month,
    MAX(energy_kwh_total) - MIN(energy_kwh_total) AS total_kwh,
    AVG(power_kw) AS avg_power_kw,
    MAX(power_kw) AS peak_power_kw,
    AVG(reactive_power_kvar) * COUNT(*) * (15.0/60.0) AS total_reactive_kvar,
    AVG(power_factor) AS avg_power_factor
FROM raw_readings
WHERE meter_id LIKE 'MM-%'
GROUP BY meter_id, date_trunc('month', timestamp)
ORDER BY meter_id, month
"""

print('  Generando meter_monthly desde raw_readings...')
start = time.time()
cur.execute(SQL_AGG)
conn.commit()
elapsed = time.time() - start
print(f'  meter_monthly insertados: {cur.rowcount} en {elapsed:.1f}s')

cur.execute("""
    SELECT COUNT(DISTINCT meter_id), COUNT(DISTINCT month), COUNT(*)
    FROM meter_monthly WHERE meter_id LIKE 'MM-%'
""")
r = cur.fetchone()
print(f'  Verificación: {r[0]} medidores, {r[1]} meses, {r[2]} filas')


# =============================================
# STEP 4: raw_readings → meter_readings (partitions)
# =============================================
print('\n' + '=' * 60)
print('STEP 4: raw_readings → meter_readings (particiones)')
print('=' * 60)

meters = sorted(store_info.keys())
conn.commit()
conn.autocommit = True

print(f'  Creando {len(meters)} particiones...')
for mid in meters:
    partition_name = f"meter_readings_{mid.lower().replace('-', '_')}"
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {partition_name}
        PARTITION OF meter_readings FOR VALUES IN ('{mid}')
    """)
print(f'  Particiones creadas: {len(meters)}')

conn.autocommit = False
total_inserted = 0
start = time.time()

for idx, mid in enumerate(meters):
    cur2 = conn.cursor(name=f'fetch_{mid}')
    cur2.itersize = BATCH_SIZE
    cur2.execute("""
        SELECT meter_id, timestamp, power_kw, reactive_power_kvar, power_factor, energy_kwh_total
        FROM raw_readings WHERE meter_id = %s ORDER BY timestamp
    """, (mid,))

    batch = []
    meter_count = 0
    for row in cur2:
        batch.append((
            row[0], row[1],
            None, None, None,  # voltage_l1, l2, l3
            None, None, None,  # current_l1, l2, l3
            row[2], row[3], row[4],  # power, reactive, pf
            None,              # frequency_hz
            row[5],            # energy_kwh_total
        ))
        if len(batch) >= BATCH_SIZE:
            args = ','.join(
                cur.mogrify("(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", b).decode()
                for b in batch
            )
            cur.execute(f"""
                INSERT INTO meter_readings
                (meter_id,timestamp,voltage_l1,voltage_l2,voltage_l3,current_l1,current_l2,current_l3,
                 power_kw,reactive_power_kvar,power_factor,frequency_hz,energy_kwh_total)
                VALUES {args}
                ON CONFLICT (meter_id, timestamp) DO NOTHING
            """)
            meter_count += len(batch)
            batch = []

    if batch:
        args = ','.join(
            cur.mogrify("(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", b).decode()
            for b in batch
        )
        cur.execute(f"""
            INSERT INTO meter_readings
            (meter_id,timestamp,voltage_l1,voltage_l2,voltage_l3,current_l1,current_l2,current_l3,
             power_kw,reactive_power_kvar,power_factor,frequency_hz,energy_kwh_total)
            VALUES {args}
            ON CONFLICT (meter_id, timestamp) DO NOTHING
        """)
        meter_count += len(batch)

    cur2.close()
    conn.commit()
    total_inserted += meter_count

    if (idx + 1) % 20 == 0 or idx == 0:
        elapsed = time.time() - start
        rate = total_inserted / elapsed if elapsed > 0 else 0
        print(f'  [{idx+1}/{len(meters)}] {mid}: {meter_count} filas | Total: {total_inserted:,} | {rate:,.0f} filas/s')

elapsed = time.time() - start
print(f'  Total meter_readings insertados: {total_inserted:,} en {elapsed:.1f}s')


# =============================================
# STEP 5: XLSX Resumen Mensual → meter_monthly_billing
# =============================================
print('\n' + '=' * 60)
print('STEP 5: XLSX → meter_monthly_billing')
print('=' * 60)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb['Resumen Mensual']
rows = list(ws.iter_rows(values_only=True))

records = []
for row in rows[3:]:
    mes_name = row[0]
    if not mes_name or mes_name not in MONTH_MAP:
        continue
    records.append((
        row[3],                                  # meter_id
        date(YEAR, MONTH_MAP[mes_name], 1),      # month
        BUILDING,
        row[7],   # total_kwh
        row[12],  # energia_clp
        row[13],  # dda_max_kw
        row[14],  # dda_max_punta_kw
        row[15],  # kwh_troncal
        row[16],  # kwh_serv_publico
        row[17],  # cargo_fijo_clp
        row[18],  # total_neto_clp
        row[19],  # iva_clp
        row[20],  # monto_exento_clp
        row[21],  # total_con_iva_clp
    ))

print(f'  Filas parseadas: {len(records)}')

cur.execute("DELETE FROM meter_monthly_billing WHERE building_name = %s", (BUILDING,))
print(f'  Eliminados previos: {cur.rowcount}')

SQL_BILLING = """
INSERT INTO meter_monthly_billing
  (meter_id, month, building_name, total_kwh, energia_clp, dda_max_kw, dda_max_punta_kw,
   kwh_troncal, kwh_serv_publico, cargo_fijo_clp, total_neto_clp,
   iva_clp, monto_exento_clp, total_con_iva_clp)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""
cur.executemany(SQL_BILLING, records)
conn.commit()
print(f'  meter_monthly_billing insertados: {cur.rowcount}')


# =============================================
# STEP 6: XLSX Resumen Mensual → KPIs (update)
# =============================================
print('\n' + '=' * 60)
print('STEP 6: XLSX → KPI columns (update)')
print('=' * 60)

kpi_records = []
for row in rows[3:]:
    mes_name = row[0]
    if not mes_name or mes_name not in MONTH_MAP:
        continue
    meter_id = row[3]
    month = date(YEAR, MONTH_MAP[mes_name], 1)
    kpi_records.append((row[8], row[9], row[10], row[11], meter_id, month))

print(f'  Filas KPI parseadas: {len(kpi_records)}')

SQL_KPI = """
UPDATE meter_monthly_billing
SET peak_mensual_kw = %s,
    demanda_hora_punta_kwh = %s,
    pct_punta_consumo = %s,
    promedio_diario_kwh = %s
WHERE meter_id = %s AND month = %s
"""
updated = 0
for rec in kpi_records:
    cur.execute(SQL_KPI, rec)
    updated += cur.rowcount

conn.commit()
print(f'  KPIs actualizados: {updated}')


# =============================================
# STEP 7: XLSX Pliegos Tarifarios (Santiago) → tariff
# =============================================
print('\n' + '=' * 60)
print('STEP 7: XLSX → tariff (Santiago)')
print('=' * 60)

# Add location column if missing
cur.execute("""
    SELECT column_name FROM information_schema.columns
    WHERE table_name = 'tariff' AND column_name = 'location'
""")
if not cur.fetchone():
    # Drop PK, add location, recreate PK
    cur.execute("ALTER TABLE tariff DROP CONSTRAINT IF EXISTS tariff_pkey")
    cur.execute("ALTER TABLE tariff ADD COLUMN location TEXT NOT NULL DEFAULT 'Las Condes'")
    cur.execute("ALTER TABLE tariff ADD PRIMARY KEY (month, location)")
    conn.commit()
    print('  Columna "location" agregada a tariff, PK actualizada a (month, location)')
else:
    print('  Columna "location" ya existe')

ws_tariff = wb['Pliegos Tarifarios (Santiago)']
tariff_rows = list(ws_tariff.iter_rows(values_only=True))

tariff_records = []
for row in tariff_rows[1:]:
    mes_num = row[0]
    if not mes_num or not isinstance(mes_num, (int, float)):
        continue
    month = date(YEAR, int(mes_num), 1)
    tariff_records.append((
        month, 'Santiago',
        row[1], row[2], row[3], row[4],
        row[5], row[6], row[7], row[8], row[9],
        row[10],
    ))

print(f'  Filas tariff parseadas: {len(tariff_records)}')

cur.execute("DELETE FROM tariff WHERE location = 'Santiago'")
print(f'  Eliminados previos Santiago: {cur.rowcount}')

SQL_TARIFF = """
INSERT INTO tariff
  (month, location, consumo_energia_kwh, dda_max_suministrada_kw, dda_max_hora_punta_kw,
   kwh_sistema_troncal, kwh_serv_publico_iva1, kwh_serv_publico_iva2,
   kwh_serv_publico_iva3, kwh_serv_publico_iva4, kwh_serv_publico_iva5, cargo_fijo_clp)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""
cur.executemany(SQL_TARIFF, tariff_records)
conn.commit()
print(f'  Tariffs Santiago insertados: {len(tariff_records)}')


# =============================================
# STEP 8: Building summary (Arauco Estación)
# =============================================
print('\n' + '=' * 60)
print('STEP 8: building_summary (Arauco Estación)')
print('=' * 60)

# Aggregate from meter_monthly_billing for MM
cur.execute("""
    SELECT
        month,
        COUNT(DISTINCT meter_id) AS total_meters,
        SUM(total_kwh) AS total_kwh,
        MAX(peak_mensual_kw) AS peak_power_kw,
        SUM(demanda_hora_punta_kwh) AS demanda_punta_kwh,
        CASE WHEN SUM(total_kwh) > 0
             THEN SUM(demanda_hora_punta_kwh) / SUM(total_kwh)
             ELSE 0 END AS pct_punta,
        SUM(promedio_diario_kwh) AS promedio_diario_kwh
    FROM meter_monthly_billing
    WHERE building_name = %s
    GROUP BY month
    ORDER BY month
""", (BUILDING,))
summary_data = cur.fetchall()
print(f'  Meses con datos: {len(summary_data)}')

cur.execute("DELETE FROM building_summary WHERE building_name = %s", (BUILDING,))
print(f'  building_summary eliminados previos: {cur.rowcount}')

SQL_SUMMARY = """
INSERT INTO building_summary
  (building_name, month, total_stores, store_types, total_meters, assigned_meters, unassigned_meters,
   total_kwh, peak_power_kw, demanda_punta_kwh, pct_punta, promedio_diario_kwh)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""

summary_records = []
for row in summary_data:
    month, total_meters, total_kwh, peak, demanda, pct, prom = row
    summary_records.append((
        BUILDING, month,
        254,            # total_stores
        0,              # store_types
        254,            # total_meters
        254,            # assigned_meters
        0,              # unassigned_meters
        total_kwh,
        peak,
        demanda,
        pct,
        prom,
    ))

cur.executemany(SQL_SUMMARY, summary_records)
conn.commit()
print(f'  building_summary insertados: {len(summary_records)}')

wb.close()


# =============================================
# FINAL VERIFICATION
# =============================================
print('\n' + '=' * 60)
print('VERIFICACIÓN FINAL')
print('=' * 60)

checks = [
    ("raw_readings MM", "SELECT COUNT(*) FROM raw_readings WHERE meter_id LIKE 'MM-%'"),
    ("store MM", "SELECT COUNT(*) FROM store WHERE meter_id LIKE 'MM-%'"),
    ("store total", "SELECT COUNT(*) FROM store"),
    ("store_type total", "SELECT COUNT(*) FROM store_type"),
    ("meter_monthly MM", "SELECT COUNT(*) FROM meter_monthly WHERE meter_id LIKE 'MM-%'"),
    ("meter_monthly total", "SELECT COUNT(*) FROM meter_monthly"),
    ("meter_readings MM", "SELECT COUNT(*) FROM meter_readings WHERE meter_id LIKE 'MM-%'"),
    ("meter_readings total", "SELECT COUNT(*) FROM meter_readings"),
    ("billing MM", "SELECT COUNT(*) FROM meter_monthly_billing WHERE building_name = 'Arauco Estación'"),
    ("billing total", "SELECT COUNT(*) FROM meter_monthly_billing"),
    ("building_summary MM", "SELECT COUNT(*) FROM building_summary WHERE building_name = 'Arauco Estación'"),
    ("building_summary total", "SELECT COUNT(*) FROM building_summary"),
    ("tariff total", "SELECT COUNT(*) FROM tariff"),
    ("particiones total", "SELECT COUNT(*) FROM pg_inherits WHERE inhparent = 'meter_readings'::regclass"),
]

for label, sql in checks:
    cur.execute(sql)
    print(f'  {label}: {cur.fetchone()[0]:,}')

# Buildings overview
print('\nEdificios:')
cur.execute("""
    SELECT building_name, COUNT(*), ROUND(SUM(total_kwh)::numeric, 0)
    FROM building_summary GROUP BY building_name ORDER BY building_name
""")
for row in cur.fetchall():
    print(f'  {row[0]}: {row[1]} meses, {row[2]:,} kWh total')

cur.close()
conn.close()
print('\nDone.')
