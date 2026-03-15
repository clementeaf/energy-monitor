#!/usr/bin/env python3
"""
Ingestion of two Strip Centers into pg-arauco.

SC52 — Arauco Express Ciudad Empresarial (52 medidores) — FULL 9-step ingest
SC53 — Arauco Express El Carmen de Huechuraba (53 medidores) — CSV-only (steps 1-4, 9)
       SC53 already has billing, building_summary, and stores in DB.
       This adds raw_readings, meter_monthly, meter_readings, and recalculates KPIs.

Tariff: Huechuraba (new location, shared by both buildings)

Usage: python3 scripts/ingest-sc52-sc53-strip-centers.py
"""
import subprocess, sys, csv, time
for pkg in ['psycopg2-binary', 'openpyxl']:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

import psycopg2
import openpyxl
from datetime import date

# --- Config ---
BATCH_SIZE = 5000
YEAR = 2025

BUILDINGS = [
    {
        'prefix': 'SC52',
        'csv': 'docs/stripCenters/SC52_StripCenter_anual.csv',
        'xlsx': 'docs/stripCenters/SC52_KPIs_mensuales_2025_M.xlsx',
        'name': 'Arauco Express Ciudad Empresarial',
        'meter_count': 52,
        'full_ingest': True,  # new building, all 9 steps
    },
    {
        'prefix': 'SC53',
        'csv': 'docs/stripCenters/SC53_StripCenter_anual.csv',
        'xlsx': 'docs/stripCenters/SC53_KPIs_mensuales_2025_M.xlsx',
        'name': 'Arauco Express El Carmen de Huechuraba',
        'meter_count': 53,
        'full_ingest': False,  # existing building, CSV-only + KPI recalc
    },
]

MONTH_MAP = {
    'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
    'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
    'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12,
}

conn = psycopg2.connect(host='127.0.0.1', port=5434, dbname='arauco', user='postgres', password='arauco')
cur = conn.cursor()


def parse_decimal(val):
    if val is None or val == '':
        return None
    return float(val.replace(',', '.'))


def header(step, text):
    print(f'\n{"=" * 60}')
    print(f'{step}: {text}')
    print('=' * 60)


# =============================================
# Process each building
# =============================================
for bldg in BUILDINGS:
    PREFIX = bldg['prefix']
    CSV_PATH = bldg['csv']
    XLSX_PATH = bldg['xlsx']
    BUILDING = bldg['name']
    METER_COUNT = bldg['meter_count']
    FULL = bldg['full_ingest']
    LIKE = f'{PREFIX}-%'

    print(f'\n{"#" * 60}')
    print(f'# {BUILDING} ({PREFIX}, {METER_COUNT} medidores)')
    print(f'# Mode: {"FULL" if FULL else "CSV-only + KPI recalc"}')
    print(f'{"#" * 60}')

    # ----- STEP 1: CSV → raw_readings -----
    header('STEP 1', 'CSV → raw_readings')

    cur.execute("SELECT COUNT(*) FROM raw_readings WHERE meter_id LIKE %s", (LIKE,))
    existing = cur.fetchone()[0]
    if existing > 0:
        print(f'  Ya existen {existing:,} filas {PREFIX} en raw_readings, saltando...')
    else:
        start = time.time()
        total = 0
        batch = []

        with open(CSV_PATH, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            next(reader)
            for row in reader:
                batch.append((
                    row[0], row[1],
                    parse_decimal(row[16]), parse_decimal(row[17]),
                    parse_decimal(row[18]), parse_decimal(row[20]),
                ))
                if len(batch) >= BATCH_SIZE:
                    args = ','.join(
                        cur.mogrify("(%s,%s,%s,%s,%s,%s)", b).decode() for b in batch
                    )
                    cur.execute(f"""
                        INSERT INTO raw_readings (timestamp, meter_id, power_kw, reactive_power_kvar, power_factor, energy_kwh_total)
                        VALUES {args} ON CONFLICT DO NOTHING
                    """)
                    total += len(batch)
                    batch = []
                    if total % 500000 == 0:
                        conn.commit()
                        elapsed = time.time() - start
                        print(f'  {total:,} filas | {total/elapsed:,.0f} filas/s | {elapsed:.0f}s')

        if batch:
            args = ','.join(
                cur.mogrify("(%s,%s,%s,%s,%s,%s)", b).decode() for b in batch
            )
            cur.execute(f"""
                INSERT INTO raw_readings (timestamp, meter_id, power_kw, reactive_power_kvar, power_factor, energy_kwh_total)
                VALUES {args} ON CONFLICT DO NOTHING
            """)
            total += len(batch)

        conn.commit()
        print(f'  Total raw_readings: {total:,} en {time.time()-start:.1f}s')

    cur.execute("SELECT COUNT(*) FROM raw_readings WHERE meter_id LIKE %s", (LIKE,))
    print(f'  Verificación raw_readings {PREFIX}: {cur.fetchone()[0]:,}')

    # ----- STEP 2: CSV → store -----
    header('STEP 2', f'CSV → store ({METER_COUNT} medidores)')

    store_info = {}
    with open(CSV_PATH, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f, delimiter=';')
        next(reader)
        for row in reader:
            mid = row[1]
            if mid not in store_info:
                store_info[mid] = (row[4], row[5])
                if len(store_info) >= METER_COUNT:
                    break

    print(f'  Medidores únicos en CSV: {len(store_info)}')

    unique_types = set(st for st, _ in store_info.values())
    print(f'  Tipos de local únicos: {len(unique_types)}')

    store_type_map = {}
    cur.execute("SELECT COALESCE(MAX(id), 0) FROM store_type")
    next_id = cur.fetchone()[0] + 1

    for st_name in sorted(unique_types):
        cur.execute("SELECT id FROM store_type WHERE name = %s", (st_name,))
        row = cur.fetchone()
        if row:
            store_type_map[st_name] = row[0]
        else:
            cur.execute("INSERT INTO store_type (id, name) VALUES (%s, %s)", (next_id, st_name))
            store_type_map[st_name] = next_id
            print(f'    Nuevo store_type: "{st_name}" → id={next_id}')
            next_id += 1

    conn.commit()

    stores = []
    for mid in sorted(store_info.keys()):
        st_name, s_name = store_info[mid]
        stores.append((mid, s_name, store_type_map[st_name]))

    cur.executemany("""
        INSERT INTO store (meter_id, store_name, store_type_id) VALUES (%s, %s, %s)
        ON CONFLICT (meter_id) DO NOTHING
    """, stores)
    conn.commit()
    print(f'  Stores insertados/existentes: {len(stores)}')

    # ----- STEP 3: raw_readings → meter_monthly -----
    header('STEP 3', 'raw_readings → meter_monthly')

    cur.execute("DELETE FROM meter_monthly WHERE meter_id LIKE %s", (LIKE,))
    print(f'  meter_monthly {PREFIX} eliminados previos: {cur.rowcount}')
    conn.commit()

    start = time.time()
    cur.execute("""
        INSERT INTO meter_monthly (meter_id, month, total_kwh, avg_power_kw, peak_power_kw, total_reactive_kvar, avg_power_factor)
        SELECT meter_id, date_trunc('month', timestamp)::date AS month,
            MAX(energy_kwh_total) - MIN(energy_kwh_total),
            AVG(power_kw), MAX(power_kw),
            AVG(reactive_power_kvar) * COUNT(*) * (15.0/60.0),
            AVG(power_factor)
        FROM raw_readings WHERE meter_id LIKE %s
        GROUP BY meter_id, date_trunc('month', timestamp)
        ORDER BY meter_id, month
    """, (LIKE,))
    conn.commit()
    print(f'  meter_monthly insertados: {cur.rowcount} en {time.time()-start:.1f}s')

    cur.execute("""
        SELECT COUNT(DISTINCT meter_id), COUNT(DISTINCT month), COUNT(*)
        FROM meter_monthly WHERE meter_id LIKE %s
    """, (LIKE,))
    r = cur.fetchone()
    print(f'  Verificación: {r[0]} medidores, {r[1]} meses, {r[2]} filas')

    # ----- STEP 4: raw_readings → meter_readings (partitions) -----
    header('STEP 4', 'raw_readings → meter_readings (particiones)')

    meters = sorted(store_info.keys())
    conn.commit()
    conn.autocommit = True

    created = 0
    for mid in meters:
        partition_name = f"meter_readings_{mid.lower().replace('-', '_')}"
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {partition_name}
            PARTITION OF meter_readings FOR VALUES IN ('{mid}')
        """)
        created += 1
    print(f'  Particiones creadas/existentes: {created}')

    conn.autocommit = False
    total_inserted = 0
    start = time.time()

    for idx, mid in enumerate(meters):
        cur2 = conn.cursor(name=f'fetch_{mid}')
        cur2.itersize = BATCH_SIZE
        cur2.execute("""
            SELECT meter_id, timestamp, power_kw, reactive_power_kvar, power_factor, energy_kwh_total
            FROM raw_readings WHERE meter_id = %s ORDER BY timestamp
        """, (mid,))

        batch = []
        meter_count = 0
        for row in cur2:
            batch.append((
                row[0], row[1],
                None, None, None, None, None, None,
                row[2], row[3], row[4], None, row[5],
            ))
            if len(batch) >= BATCH_SIZE:
                args = ','.join(
                    cur.mogrify("(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", b).decode() for b in batch
                )
                cur.execute(f"""
                    INSERT INTO meter_readings
                    (meter_id,timestamp,voltage_l1,voltage_l2,voltage_l3,current_l1,current_l2,current_l3,
                     power_kw,reactive_power_kvar,power_factor,frequency_hz,energy_kwh_total)
                    VALUES {args} ON CONFLICT (meter_id, timestamp) DO NOTHING
                """)
                meter_count += len(batch)
                batch = []

        if batch:
            args = ','.join(
                cur.mogrify("(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", b).decode() for b in batch
            )
            cur.execute(f"""
                INSERT INTO meter_readings
                (meter_id,timestamp,voltage_l1,voltage_l2,voltage_l3,current_l1,current_l2,current_l3,
                 power_kw,reactive_power_kvar,power_factor,frequency_hz,energy_kwh_total)
                VALUES {args} ON CONFLICT (meter_id, timestamp) DO NOTHING
            """)
            meter_count += len(batch)

        cur2.close()
        conn.commit()
        total_inserted += meter_count

        if (idx + 1) % 10 == 0 or idx == 0:
            elapsed = time.time() - start
            print(f'  [{idx+1}/{len(meters)}] {mid}: {meter_count} | Total: {total_inserted:,} | {total_inserted/elapsed:,.0f} filas/s')

    print(f'  Total meter_readings: {total_inserted:,} en {time.time()-start:.1f}s')

    # ----- STEPS 5-8: XLSX billing + building_summary (SC52 only) -----
    if FULL:
        wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

        # STEP 5: billing
        header('STEP 5', 'XLSX → meter_monthly_billing')
        ws = wb['Resumen Mensual']
        rows_data = list(ws.iter_rows(values_only=True))

        records = []
        for row in rows_data[3:]:
            mes_name = row[0]
            if not mes_name or mes_name not in MONTH_MAP:
                continue
            records.append((
                row[3], date(YEAR, MONTH_MAP[mes_name], 1), BUILDING,
                row[7], row[12], row[13], row[14], row[15], row[16],
                row[17], row[18], row[19], row[20], row[21],
            ))

        print(f'  Filas parseadas: {len(records)}')
        cur.execute("DELETE FROM meter_monthly_billing WHERE building_name = %s", (BUILDING,))
        print(f'  Eliminados previos: {cur.rowcount}')

        cur.executemany("""
            INSERT INTO meter_monthly_billing
              (meter_id, month, building_name, total_kwh, energia_clp, dda_max_kw, dda_max_punta_kw,
               kwh_troncal, kwh_serv_publico, cargo_fijo_clp, total_neto_clp,
               iva_clp, monto_exento_clp, total_con_iva_clp)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, records)
        conn.commit()
        print(f'  meter_monthly_billing insertados: {cur.rowcount}')

        # STEP 6: KPIs
        header('STEP 6', 'XLSX → KPI columns (update)')
        kpi_records = []
        for row in rows_data[3:]:
            mes_name = row[0]
            if not mes_name or mes_name not in MONTH_MAP:
                continue
            kpi_records.append((row[8], row[9], row[10], row[11], row[3], date(YEAR, MONTH_MAP[mes_name], 1)))

        print(f'  Filas KPI: {len(kpi_records)}')
        updated = 0
        for rec in kpi_records:
            cur.execute("""
                UPDATE meter_monthly_billing
                SET peak_mensual_kw = %s, demanda_hora_punta_kwh = %s,
                    pct_punta_consumo = %s, promedio_diario_kwh = %s
                WHERE meter_id = %s AND month = %s
            """, rec)
            updated += cur.rowcount
        conn.commit()
        print(f'  KPIs actualizados: {updated}')

        # STEP 8: building_summary
        header('STEP 8', f'building_summary ({BUILDING})')
        cur.execute("""
            SELECT month, COUNT(DISTINCT meter_id),
                SUM(total_kwh), MAX(peak_mensual_kw),
                SUM(demanda_hora_punta_kwh),
                CASE WHEN SUM(total_kwh) > 0 THEN SUM(demanda_hora_punta_kwh)/SUM(total_kwh) ELSE 0 END,
                SUM(promedio_diario_kwh)
            FROM meter_monthly_billing WHERE building_name = %s
            GROUP BY month ORDER BY month
        """, (BUILDING,))
        summary_data = cur.fetchall()
        print(f'  Meses: {len(summary_data)}')

        cur.execute("DELETE FROM building_summary WHERE building_name = %s", (BUILDING,))
        print(f'  Eliminados previos: {cur.rowcount}')

        for row in summary_data:
            month, total_meters, total_kwh, peak, demanda, pct, prom = row
            cur.execute("""
                INSERT INTO building_summary
                  (building_name, month, total_stores, store_types, total_meters, assigned_meters, unassigned_meters,
                   total_kwh, peak_power_kw, demanda_punta_kwh, pct_punta, promedio_diario_kwh)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (BUILDING, month, METER_COUNT, 0, METER_COUNT, METER_COUNT, 0,
                  total_kwh, peak, demanda, pct, prom))
        conn.commit()
        print(f'  building_summary insertados: {len(summary_data)}')

        wb.close()

    # ----- STEP 9: building_summary KPIs -----
    header('STEP 9', 'building_summary KPIs (CRITICAL)')

    cur.execute("""
        UPDATE building_summary bs SET
            peak_demand_kw   = sub.sum_peak,
            avg_power_kw     = sub.avg_pw,
            avg_power_factor = sub.avg_pf
        FROM (
            SELECT mm.month,
                SUM(mm.peak_power_kw)    AS sum_peak,
                AVG(mm.avg_power_kw)     AS avg_pw,
                AVG(mm.avg_power_factor) AS avg_pf
            FROM meter_monthly mm
            WHERE mm.meter_id LIKE %s
            GROUP BY mm.month
        ) sub
        WHERE bs.building_name = %s
          AND bs.month = sub.month - INTERVAL '1 year'
    """, (LIKE, BUILDING))
    print(f'  KPIs actualizados: {cur.rowcount}')
    conn.commit()

    cur.execute("""
        SELECT COUNT(*), COUNT(peak_demand_kw), COUNT(avg_power_kw), COUNT(avg_power_factor)
        FROM building_summary WHERE building_name = %s
    """, (BUILDING,))
    v = cur.fetchone()
    print(f'  Verificación: {v[0]} meses, peak={v[1]}, avg_kw={v[2]}, avg_pf={v[3]}')
    if v[0] == v[1] == v[2] == v[3]:
        print(f'  OK — {PREFIX} KPIs completos')
    else:
        print(f'  ALERTA: {PREFIX} tiene KPIs NULL!')


# =============================================
# STEP 7: Tariff Huechuraba (once, from SC52 XLSX)
# =============================================
header('STEP 7', 'XLSX → tariff (Huechuraba)')

wb = openpyxl.load_workbook(BUILDINGS[0]['xlsx'], read_only=True, data_only=True)
ws_tariff = wb['Pliegos Tarifarios (Huechuraba)']
tariff_rows = list(ws_tariff.iter_rows(values_only=True))

tariff_records = []
for row in tariff_rows[1:]:
    mes_num = row[0]
    if not mes_num or not isinstance(mes_num, (int, float)):
        continue
    tariff_records.append((
        date(YEAR, int(mes_num), 1), 'Huechuraba',
        row[1], row[2], row[3], row[4],
        row[5], row[6], row[7], row[8], row[9], row[10],
    ))

print(f'  Filas tariff parseadas: {len(tariff_records)}')
cur.execute("DELETE FROM tariff WHERE location = 'Huechuraba'")
print(f'  Eliminados previos: {cur.rowcount}')

cur.executemany("""
    INSERT INTO tariff
      (month, location, consumo_energia_kwh, dda_max_suministrada_kw, dda_max_hora_punta_kw,
       kwh_sistema_troncal, kwh_serv_publico_iva1, kwh_serv_publico_iva2,
       kwh_serv_publico_iva3, kwh_serv_publico_iva4, kwh_serv_publico_iva5, cargo_fijo_clp)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
""", tariff_records)
conn.commit()
print(f'  Tariffs Huechuraba insertados: {len(tariff_records)}')
wb.close()


# =============================================
# FINAL VERIFICATION
# =============================================
header('FINAL', 'VERIFICACIÓN')

checks = [
    ("raw_readings SC52", "SELECT COUNT(*) FROM raw_readings WHERE meter_id LIKE 'SC52-%%'"),
    ("raw_readings SC53", "SELECT COUNT(*) FROM raw_readings WHERE meter_id LIKE 'SC53-%%'"),
    ("store total", "SELECT COUNT(*) FROM store"),
    ("store_type total", "SELECT COUNT(*) FROM store_type"),
    ("meter_monthly total", "SELECT COUNT(*) FROM meter_monthly"),
    ("meter_readings total", "SELECT COUNT(*) FROM meter_readings"),
    ("billing total", "SELECT COUNT(*) FROM meter_monthly_billing"),
    ("building_summary total", "SELECT COUNT(*) FROM building_summary"),
    ("tariff total", "SELECT COUNT(*) FROM tariff"),
    ("particiones total", "SELECT COUNT(*) FROM pg_inherits WHERE inhparent = 'meter_readings'::regclass"),
]

for label, sql in checks:
    cur.execute(sql)
    print(f'  {label}: {cur.fetchone()[0]:,}')

print('\nEdificios:')
cur.execute("""
    SELECT building_name, COUNT(*), ROUND(SUM(total_kwh)::numeric, 0),
           COUNT(peak_demand_kw), COUNT(avg_power_kw), COUNT(avg_power_factor)
    FROM building_summary GROUP BY building_name ORDER BY building_name
""")
for row in cur.fetchall():
    print(f'  {row[0]}: {row[1]} meses, {row[2]:,} kWh, KPIs: peak={row[3]} avg={row[4]} pf={row[5]}')

cur.close()
conn.close()
print('\nDone.')
