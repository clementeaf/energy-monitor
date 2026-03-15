#!/usr/bin/env python3
"""Script 5: Load SC53 Arauco Express data from XLSX (Consumo por Local + Resumen Ejecutivo)."""
import subprocess, sys
for pkg in ['openpyxl', 'psycopg2-binary']:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

import openpyxl
import psycopg2
from datetime import date

XLSX = 'docs/porCargar/MG446_KPIs_mensuales_2025_M.xlsx'
BUILDING = 'Arauco Express El Carmen de Huechuraba'
YEAR = 2025

MONTH_MAP = {
    'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
    'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
    'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12,
}

conn = psycopg2.connect(host='127.0.0.1', port=5434, dbname='arauco', user='postgres', password='arauco')
cur = conn.cursor()

# =============================================
# PART 1: Consumo por Local (Pivot) → meter_monthly_billing
# =============================================
print('=== Consumo por Local (Pivot) ===')
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb['Consumo por Local (Pivot)']
rows = list(ws.iter_rows(values_only=True))

# 3 sections: ① kWh (rows 4-56), ② Peak kW (rows 61-113), ③ Demanda Punta (rows 118-170)
# Parse each section into dicts keyed by (meter_id, month)
kwh_data = {}   # (meter_id, month) → total_kwh
peak_data = {}  # (meter_id, month) → peak_kw
demand_data = {} # (meter_id, month) → demanda_punta_kwh

sections = [(4, 57, kwh_data), (61, 114, peak_data), (118, 171, demand_data)]
for start, end, target in sections:
    for row in rows[start:end]:
        meter_id = row[0]
        if not meter_id or not isinstance(meter_id, str) or not meter_id.startswith('SC53'):
            continue
        for month_idx in range(12):
            val = row[2 + month_idx]
            if val is None:
                continue
            target[(meter_id, date(YEAR, month_idx + 1, 1))] = val

# Build billing records with all 3 metrics
all_keys = sorted(set(kwh_data.keys()) | set(peak_data.keys()) | set(demand_data.keys()))
billing_records = []
for key in all_keys:
    meter_id, month = key
    billing_records.append((
        meter_id, month, BUILDING,
        kwh_data.get(key),
        peak_data.get(key),
        demand_data.get(key),
    ))

print(f'Filas parseadas: {len(billing_records)}')

# Delete existing SC53 billing data
cur.execute("DELETE FROM meter_monthly_billing WHERE building_name = %s", (BUILDING,))
print(f'Eliminados previos: {cur.rowcount}')

# Insert with total_kwh, peak_mensual_kw, demanda_hora_punta_kwh
SQL_BILLING = """
INSERT INTO meter_monthly_billing (meter_id, month, building_name, total_kwh, peak_mensual_kw, demanda_hora_punta_kwh)
VALUES (%s, %s, %s, %s, %s, %s)
"""
cur.executemany(SQL_BILLING, billing_records)
conn.commit()
print(f'meter_monthly_billing insertados: {len(billing_records)}')

# =============================================
# PART 2: Resumen Ejecutivo → building_summary
# =============================================
print('\n=== Resumen Ejecutivo ===')
ws2 = wb['Resumen Ejecutivo']
rows2 = list(ws2.iter_rows(values_only=True))

# Row 2 header: Mes, N°Mes, Consumo Total, Peak Máx, Demanda Punta Total, %Punta, Promedio Diario, Local Mayor Consumo
# Rows 3..14 data (12 months)
summary_records = []
for row in rows2[3:]:
    mes_name = row[0]
    if not mes_name or mes_name not in MONTH_MAP:
        continue
    month = date(YEAR, MONTH_MAP[mes_name], 1)
    summary_records.append((
        BUILDING,
        month,
        53,     # total_stores (53 locales SC53)
        0,      # store_types
        53,     # total_meters
        53,     # assigned_meters
        0,      # unassigned_meters
        row[2], # total_kwh (Consumo Total Centro)
        row[3], # peak_power_kw (Peak Máx Centro)
        row[4], # demanda_punta_kwh (Demanda Punta Total)
        row[5], # pct_punta (% Punta)
        row[6], # promedio_diario_kwh (Promedio Diario Centro)
    ))

wb.close()
print(f'Filas resumen parseadas: {len(summary_records)}')

# Check building_summary structure — add columns if needed
cur.execute("""
    SELECT column_name FROM information_schema.columns
    WHERE table_name = 'building_summary' ORDER BY ordinal_position
""")
existing_cols = [r[0] for r in cur.fetchall()]
print(f'Columnas building_summary: {existing_cols}')

# Add SC53-specific columns if missing
for col_def in [
    'peak_power_kw NUMERIC(12,3)',
    'demanda_punta_kwh NUMERIC(14,3)',
    'pct_punta NUMERIC(5,4)',
    'promedio_diario_kwh NUMERIC(12,3)',
]:
    col_name = col_def.split()[0]
    if col_name not in existing_cols:
        cur.execute(f"ALTER TABLE building_summary ADD COLUMN {col_def}")
        print(f'  Columna {col_name} agregada')
conn.commit()

# Delete existing SC53 summary
cur.execute("DELETE FROM building_summary WHERE building_name = %s", (BUILDING,))
print(f'building_summary eliminados previos: {cur.rowcount}')

# Insert — use available columns
SQL_SUMMARY = """
INSERT INTO building_summary
  (building_name, month, total_stores, store_types, total_meters, assigned_meters, unassigned_meters,
   total_kwh, peak_power_kw, demanda_punta_kwh, pct_punta, promedio_diario_kwh)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""
cur.executemany(SQL_SUMMARY, summary_records)
conn.commit()
print(f'building_summary insertados: {len(summary_records)}')

# =============================================
# Verification
# =============================================
print('\n=== Verificación ===')
cur.execute("SELECT COUNT(*) FROM meter_monthly_billing WHERE building_name = %s", (BUILDING,))
print(f'meter_monthly_billing SC53: {cur.fetchone()[0]} filas')

cur.execute("SELECT COUNT(*) FROM building_summary WHERE building_name = %s", (BUILDING,))
print(f'building_summary SC53: {cur.fetchone()[0]} filas')

cur.execute("""
    SELECT month, total_kwh, peak_power_kw, promedio_diario_kwh
    FROM building_summary WHERE building_name = %s ORDER BY month
""", (BUILDING,))
print('\nResumen SC53 por mes:')
for row in cur.fetchall():
    print(f'  {row[0]}: {row[1]:,.1f} kWh, peak={row[2]} kW, prom_diario={row[3]:,.1f} kWh')

cur.close()
conn.close()
print('\nDone.')
